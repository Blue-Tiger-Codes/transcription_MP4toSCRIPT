import speech_recognition as sr
from moviepy.editor import VideoFileClip
import openpyxl
from openpyxl.styles import Alignment

def create_timestamp_spreadsheet(mp4_file, output_file):
    # Load the MP4 file
    video = VideoFileClip(mp4_file)

    # Get the duration of the video
    duration = video.duration

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add the header row
    worksheet['A1'] = 'Timestamp'
    worksheet['B1'] = 'Transcript'
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet['B1'].alignment = Alignment(horizontal='center')

    # Initialize speech recognizer
    r = sr.Recognizer()

    # Process the video in 5-second chunks
    for i in range(0, int(duration), 5):
        start_time = i
        end_time = min(i + 5, duration)

        # Extract audio for this chunk
        audio_chunk = video.subclip(start_time, end_time).audio
        audio_chunk.write_audiofile("temp_audio.wav", codec='pcm_s16le')

        # Transcribe audio
        with sr.AudioFile("temp_audio.wav") as source:
            audio = r.record(source)
            try:
                text = r.recognize_google(audio, language="en-US")
            except sr.UnknownValueError:
                text = "Could not understand audio"
            except sr.RequestError:
                text = "Could not request results"

        # Add timestamp and transcript to worksheet
        timestamp = f"{int(start_time // 3600):02d}:{int((start_time % 3600) // 60):02d}:{int(start_time % 60):02d}"
        row = i // 5 + 2
        worksheet.cell(row=row, column=1, value=timestamp)
        worksheet.cell(row=row, column=2, value=text)

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 50

    # Save the workbook
    workbook.save(output_file)

    # Clean up
    video.close()

    print(f"Timestamp spreadsheet created: {output_file}")

# Example usage
create_timestamp_spreadsheet(r"C:\Users\rajmy\OneDrive\Desktop\project_hp_ recon\palki.mp4", 'output_timestamps.xlsx')
# Instead of "C:\Users\rajmy\OneDrive\Desktop\project_hp_ recon\palki.mp4"use you file path